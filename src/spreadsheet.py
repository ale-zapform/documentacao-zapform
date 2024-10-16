import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

def gerar_planilha(data_geral, data_por_cliente, reference_month):
    """
    Gera uma planilha Excel com múltiplas abas:
    - A aba "Geral" com todos os dados agregados.
    - Abas separadas para cada ZF Client com dados por configuração.
    
    :param data_geral: Lista com os dados para a aba geral.
    :param data_por_cliente: Dicionário com os dados organizados por ZF Client.
    :param reference_month: Mês de referência para o nome do arquivo.
    """
    # Criar o arquivo Excel
    with pd.ExcelWriter(f"faturamento_zapform_{reference_month}.xlsx", engine='openpyxl') as writer:
        # Aba geral com todas as informações
        df_geral = pd.DataFrame(data_geral, columns=['ZF Client', 'Config ID', 'Config Name', 'Status Code', 'Status Name', 'Order Count'])
        df_geral.to_excel(writer, sheet_name='Geral', index=False)

        # Estilos
        bold_font = Font(name='Calibri', size=11, bold=True)
        normal_font = Font(name='Calibri', size=11)
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        center_align = Alignment(horizontal='center')
        left_align = Alignment(horizontal='left')

        # Formatando a aba Geral
        wb = writer.book
        ws_geral = wb['Geral']
        
        # Definir as larguras das colunas na aba "Geral"
        ws_geral.column_dimensions['A'].width = 12
        ws_geral.column_dimensions['B'].width = 12
        ws_geral.column_dimensions['C'].width = 20
        ws_geral.column_dimensions['D'].width = 12
        ws_geral.column_dimensions['E'].width = 40
        ws_geral.column_dimensions['F'].width = 12

        # Aplicar fundo cinza claro e centralizar o cabeçalho da aba "Geral"
        for cell in ws_geral[1]:
            cell.fill = gray_fill
            cell.font = bold_font
            cell.alignment = center_align

        # Centralizar todos os dados na aba "Geral", exceto colunas C e E
        for row in ws_geral.iter_rows(min_row=2, max_row=ws_geral.max_row, min_col=1, max_col=6):
            for cell in row:
                if cell.column_letter in ['C', 'E']:  # Colunas C e E alinhadas à esquerda
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.font = normal_font

        # Criar abas para cada cliente
        for zfclient_id, data_cliente in data_por_cliente.items():
            # Criar uma nova planilha para o cliente
            ws = wb.create_sheet(title=f'ZF Client {zfclient_id}')
            
            # Definir as larguras das colunas A, B e C
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 12

            row_num = 1  # Iniciar no topo da planilha
            total_ordens_geral = 0  # Totalizador geral de todas as configurações

            # Loop por todas as configurações e dados do cliente
            for data in data_cliente:
                if "Config:" in data[0]:  # Se for a linha de título de uma configuração
                    # Antes de mudar para a próxima configuração, adicione a linha em branco
                    if row_num > 1:
                        row_num += 1  # Pula uma linha entre as configurações

                    # O total de ordens para a configuração já está calculado em `data[2]`
                    total_ordens_config = data[2]

                    # Garantir que `total_ordens_config` é um número antes de somar
                    if isinstance(total_ordens_config, (int, float)):
                        total_ordens_geral += total_ordens_config  # Acumular no totalizador geral

                    # Título da Configuração
                    ws.cell(row=row_num, column=1).value = f"{data[0]} ({data[1]})"
                    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='left')
                    ws.cell(row=row_num, column=1).font = bold_font

                    # Número total de ordens da configuração na coluna C
                    ws.cell(row=row_num, column=3).value = total_ordens_config
                    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
                    ws.cell(row=row_num, column=3).font = bold_font

                    row_num += 1

                    # Adicionar cabeçalho da tabela de status em negrito com fundo cinza claro
                    headers = ['Status Code', 'Status Name', 'Order Count']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = header
                        cell.font = bold_font
                        cell.fill = gray_fill
                        cell.alignment = center_align
                    row_num += 1
                else:
                    # Adicionar os dados dos status
                    for col_num, cell_value in enumerate(data[:3], 1):
                        ws.cell(row=row_num, column=col_num).value = cell_value
                        ws.cell(row=row_num, column=col_num).font = normal_font
                        ws.cell(row=row_num, column=col_num).alignment = center_align
                    row_num += 1

    print(f"\nPlanilha gerada: faturamento_zapform_{reference_month}.xlsx")
